"""
Report Generator
Generates Excel and CSV reports from ERP data
"""

import logging
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generate reports from ERP data"""
    
    def __init__(self, report_dir: str = 'reports'):
        """
        Initialize report generator
        
        Args:
            report_dir: Directory to save reports
        """
        self.report_dir = Path(report_dir)
        self.csv_dir = self.report_dir / 'csv'
        self._ensure_directories()
        logger.info("ReportGenerator initialized")
    
    def _ensure_directories(self):
        """Create report directories if they don't exist"""
        self.report_dir.mkdir(exist_ok=True)
        self.csv_dir.mkdir(exist_ok=True)
    
    def _format_excel_sheet(self, ws, df: pd.DataFrame):
        """
        Apply formatting to Excel sheet
        
        Args:
            ws: Worksheet object
            df: DataFrame to format
        """
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for col_num, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Format data rows
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border
                
                # Alternate row colors
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Auto-adjust column widths
        for col_num, column in enumerate(df.columns, 1):
            max_length = max(
                df[column].astype(str).apply(len).max(),
                len(column)
            )
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def generate_excel_reports(
        self,
        purchase_df: pd.DataFrame,
        production_df: pd.DataFrame,
        packing_df: pd.DataFrame,
        shipment_df: pd.DataFrame
    ):
        """
        Generate comprehensive Excel report with multiple sheets
        
        Args:
            purchase_df: Purchase data
            production_df: Production data
            packing_df: Packing data
            shipment_df: Shipment data
        """
        logger.info("Generating Excel reports...")
        
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = self.report_dir / f'daily_report_{timestamp}.xlsx'
            latest_filename = self.report_dir / 'daily_report.xlsx'
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Write each department data to separate sheets
                purchase_df.to_excel(writer, sheet_name='Purchase', index=False)
                production_df.to_excel(writer, sheet_name='Production', index=False)
                packing_df.to_excel(writer, sheet_name='Packing', index=False)
                shipment_df.to_excel(writer, sheet_name='Shipment', index=False)
                
                # Create summary sheet
                summary_data = self._create_summary(
                    purchase_df, production_df, packing_df, shipment_df
                )
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Apply formatting
                workbook = writer.book
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    if sheet_name == 'Summary':
                        self._format_summary_sheet(ws, summary_df)
                    else:
                        # Get the dataframe for this sheet
                        df_map = {
                            'Purchase': purchase_df,
                            'Production': production_df,
                            'Packing': packing_df,
                            'Shipment': shipment_df
                        }
                        if sheet_name in df_map:
                            self._format_excel_sheet(ws, df_map[sheet_name])
            
            # Save a copy as latest report
            import shutil
            shutil.copy(filename, latest_filename)
            
            logger.info(f"Excel report generated: {filename}")
            
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {str(e)}", exc_info=True)
            raise
    
    def _create_summary(
        self,
        purchase_df: pd.DataFrame,
        production_df: pd.DataFrame,
        packing_df: pd.DataFrame,
        shipment_df: pd.DataFrame
    ) -> List[Dict]:
        """Create summary statistics"""
        summary = [
            {
                'Department': 'Purchase',
                'Total Records': len(purchase_df),
                'Pending': len(purchase_df[purchase_df['status'] == 'Pending']),
                'Completed': len(purchase_df[purchase_df['status'].isin(['Approved', 'Delivered'])]),
                'Total Value': f"₹{purchase_df['amount'].sum():,.2f}" if 'amount' in purchase_df else 'N/A'
            },
            {
                'Department': 'Production',
                'Total Records': len(production_df),
                'Pending': len(production_df[production_df['status'] == 'Pending']),
                'Completed': len(production_df[production_df['status'] == 'Completed']),
                'Total Quantity': f"{production_df['quantity'].sum():,}" if 'quantity' in production_df else 'N/A'
            },
            {
                'Department': 'Packing',
                'Total Records': len(packing_df),
                'Pending': len(packing_df[packing_df['status'] == 'Pending']),
                'Completed': len(packing_df[packing_df['status'] == 'Completed']),
                'Total Packed': f"{packing_df['quantity'].sum():,}" if 'quantity' in packing_df else 'N/A'
            },
            {
                'Department': 'Shipment',
                'Total Records': len(shipment_df),
                'Pending': len(shipment_df[shipment_df['status'] == 'Pending']),
                'Completed': len(shipment_df[shipment_df['status'] == 'Delivered']),
                'Total Shipped': f"{shipment_df['quantity'].sum():,}" if 'quantity' in shipment_df else 'N/A'
            }
        ]
        return summary
    
    def _format_summary_sheet(self, ws, df: pd.DataFrame):
        """Format the summary sheet with enhanced styling"""
        self._format_excel_sheet(ws, df)
        
        # Add title
        ws.insert_rows(1)
        ws['A1'] = f"ERP Daily Summary Report - {datetime.now().strftime('%Y-%m-%d')}"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:F1')
    
    def generate_csv_exports(
        self,
        purchase_df: pd.DataFrame,
        production_df: pd.DataFrame,
        packing_df: pd.DataFrame,
        shipment_df: pd.DataFrame
    ):
        """
        Generate CSV exports for each department
        
        Args:
            purchase_df: Purchase data
            production_df: Production data
            packing_df: Packing data
            shipment_df: Shipment data
        """
        logger.info("Generating CSV exports...")
        
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Export each dataset
            datasets = {
                'purchase_data': purchase_df,
                'production_data': production_df,
                'packing_data': packing_df,
                'shipment_data': shipment_df
            }
            
            for name, df in datasets.items():
                filename = self.csv_dir / f'{name}_{timestamp}.csv'
                latest_filename = self.csv_dir / f'{name}.csv'
                
                df.to_csv(filename, index=False)
                df.to_csv(latest_filename, index=False)
                
                logger.info(f"CSV exported: {filename}")
            
        except Exception as e:
            logger.error(f"Failed to generate CSV exports: {str(e)}", exc_info=True)
            raise
    
    def generate_department_summary(self, df: pd.DataFrame, department: str) -> pd.DataFrame:
        """
        Generate summary for a specific department
        
        Args:
            df: Department data
            department: Department name
            
        Returns:
            Summary DataFrame
        """
        if 'status' not in df.columns:
            return pd.DataFrame()
        
        summary = df.groupby('status').agg({
            df.columns[0]: 'count'
        }).reset_index()
        summary.columns = ['Status', 'Count']
        summary['Department'] = department
        
        return summary
    
    def cleanup_old_reports(self, days: int = 30):
        """
        Clean up old report files
        
        Args:
            days: Keep reports newer than this many days
        """
        logger.info(f"Cleaning up reports older than {days} days...")
        
        try:
            cutoff_time = datetime.now().timestamp() - (days * 86400)
            
            for file_path in self.report_dir.rglob('*_*.xlsx'):
                if file_path.stat().st_mtime < cutoff_time:
                    file_path.unlink()
                    logger.info(f"Deleted old report: {file_path}")
            
            for file_path in self.csv_dir.rglob('*_*.csv'):
                if file_path.stat().st_mtime < cutoff_time:
                    file_path.unlink()
                    logger.info(f"Deleted old CSV: {file_path}")
                    
        except Exception as e:
            logger.error(f"Error during cleanup: {str(e)}", exc_info=True)
